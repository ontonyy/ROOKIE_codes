import openpyxl
import pandas as pd

data = pd.read_csv("Digiehituse_andmestik_2023.csv")
print(data.columns)

# Load the Excel file
workbook = openpyxl.load_workbook('Digiehituse_andmestik_2023.xlsx')

print(workbook.sheetnames)

# Select the worksheet
worksheet = workbook['Digiehituse_andmestik']  # Replace 'Sheet1' with the actual name of the worksheet

columns = worksheet.iter_cols()
for col in columns:
    print(col[0].column_letter)

prev = None
for cell in worksheet['AW']:
    if cell.value is None:
        print(prev)
        print("Found an empty cell in column")
        break
    prev = cell.value
else:
    print("All cells in column A have a value.")